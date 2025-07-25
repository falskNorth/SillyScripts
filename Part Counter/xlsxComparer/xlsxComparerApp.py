import os
from openpyxl import load_workbook
from collections import defaultdict
import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD


def count_duplicates(folder_path, output_filename):
    item_counts = defaultdict(int)
    quantity_counts = defaultdict(str)  # STORE QUANTITY'S AS STRINGS
    ar_items = []  # LIST FOR 'A/R' ITEMS

    # LIST ALL .xlsx FILES IN THE DIR
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            print(f"Processing file: {filename}")

            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active  # GET THE FIRST SHEET

                # ITERATE THROUGH THE ROWS, SKIPPING THE FIRST 10 ROWS
                for row in sheet.iter_rows(min_row=11, values_only=True):
                    item = row[7]  # H COLUMN IS INDEX 7 (0-BASED INDEX)
                    quantity = row[3]  # D COLUMN IS INDEX 3 (0-BASED INDEX)
                    status = row[3]  # D COLUMN IS INDEX 3 (0-BASED INDEX) FOR 'A/R'

                    if item is not None:
                        item = str(item).strip()  # CONVERT TO STRING AND TRIM WHITESPACE
                        if item and item != '0':  # IGNORE '0'
                            item_counts[item] += 1  # COUNT OCCURRENCES

                            if status == 'A/R':  # CHECK IF STATUS OF STR IS 'A/R'
                                ar_items.append(item)  # ADD TO 'A/R' LIST
                                quantity_counts[item] = 'A/R'  # LABLE QUANTITY AS 'A/R'
                            else:
                                # CHECK IF QUANTITY IS VALID INT
                                if quantity is not None:
                                    quantity_str = str(quantity).strip()  # CONVERT TO STRING AND TRIM WHITESPACE
                                    if quantity_str.isdigit():  # CHECK IF INT
                                        # IF INT, SUM IT AS AN INT AND CONVERT TO STR
                                        if quantity_counts[item] == '':  # IF FIRST ENTRY INITIALISE
                                            quantity_counts[item] = quantity_str
                                        else:
                                            quantity_counts[item] = str(
                                                int(quantity_counts[item]) + int(quantity_str))  # SUM QUANTITIES

            except Exception as e:
                print(f"Error processing file {filename}: {e}")

    # PREPARE TO WRITE RESULT TO .txt FILE IN OUTPUT FOLDER
    output_folder = os.path.join(folder_path, 'Output')
    os.makedirs(output_folder, exist_ok=True)  # CREATE 'Output' FOLDER IF NONE EXIST

    # AUTO ADD .txt EXTENSION
    if not output_filename.endswith('.txt'):
        output_filename += '.txt'

    output_file_path = os.path.join(output_folder, output_filename)

    # OPEN OUTPUT FILE IN WRITE MODE TO SAVE RESULTS
    with open(output_file_path, 'w') as output_file:
        # PRINT SORTED RESULTS ALPHABETICALLY
        if not item_counts:
            print("No items found.")
            output_file.write("No items found.\n")  # WRITE TO TXT FILE
        else:
            for item in sorted(item_counts.keys()):  # SORT ITEMS ALPHABETICALLY
                count = item_counts[item]
                total_quantity = quantity_counts[item]

                # PRINT 'A/R' IN TQ IF APPLICABLE
                if total_quantity == 'A/R':
                    output_line = f"{item}: Count = {count}, Total Quantity = A/R"
                else:
                    output_line = f"{item}: Count = {count}, Total Quantity = {total_quantity}"

                print(output_line)
                output_file.write(output_line + '\n')  # WRITE TO TXT FILE

            # WRITE A/R TIMES TO OUTPUT FILE
            if ar_items:
                output_file.write("\nItems with 'A/R':\n")
                for ar_item in sorted(set(ar_items)):  # USE SET TO AVOID DUPS
                    output_file.write(f"{ar_item}\n")
                print("\nItems with 'A/R':")
                for ar_item in sorted(set(ar_items)):
                    print(ar_item)

    print(f"Results saved to  {output_file_path}")


def drop(event):
    # GET FILE PATH FROM DROP EVENT
    file_path = event.data
    if file_path.endswith('.xlsx'):
        # MOVE FILE TO 'Docs' FOLDER
        target_folder = "Docs/"
        os.makedirs(target_folder, exist_ok=True)
        target_path = os.path.join(target_folder, os.path.basename(file_path))
        os.rename(file_path, target_path)
        print(f"File moved to {target_path}")

        # ASKS FOR OUTPUT FILENAME OUT COUNT DUPLICATES
        output_filename = output_entry.get()
        count_duplicates(target_folder, output_filename)


# MAIN WINDOW
root = TkinterDnD.Tk()
root.title("Drag and Drop Excel Files")
root.geometry("400x200")

# CREATE INSTRUCTIONS LABEL
label = tk.Label(root, text="Drag and drop .xlsx files here:")
label.pack(pady=20)

# CREATE ENTRY FOR OUTPUT FILENAME
output_entry = tk.Entry(root, width=30)
output_entry.pack(pady=10)
output_entry.insert(0, "Input Filename Here")  # DEFULT OUTPUT FILE NAME

# CREATE DROP TARGET
root.drop_target_register(DND_FILES)
root.dnd_bind('<<Drop>>', drop)

# GUI LOOP START
root.mainloop()
